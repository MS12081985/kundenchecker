import pandas as pd
import pytest
from openpyxl import load_workbook
from PySide6.QtCore import Qt

from excel.importer import load_excel
from models.address_utils import normalize_postal_code, normalize_street, street_match
from services.bulk_research_service import BulkResearchService
from services.customer_export_service import CustomerExportService
from services.import_quality_service import ImportQualityService
from services.research_service import ResearchService
from ui.table_model import CustomerTableModel


@pytest.mark.parametrize("value, expected", [
    (61462.0, "61462"), ("61462.0", "61462"), (60313, "60313"),
    ("01067", "01067"), (None, ""), (float("nan"), ""), (61462.5, "61462.5"),
])
def test_postal_code_normalization(value, expected):
    assert normalize_postal_code(value) == expected


def test_postal_code_import_table_and_exports(tmp_path):
    source = tmp_path / "source.xlsx"
    pd.DataFrame({"KUNDENNAME": ["A", "B"], "PLZ": ["01067", "61462.0"]}).to_excel(source, index=False)
    frame = load_excel(str(source))
    assert frame["PLZ"].tolist() == ["01067", "61462"]
    model = CustomerTableModel(pd.DataFrame({"PLZ": [61462.0]}))
    assert model.data(model.index(0, 0), Qt.DisplayRole) == "61462"

    xlsx = CustomerExportService().write(frame, tmp_path / "export.xlsx", "xlsx")
    sheet = load_workbook(xlsx).active
    assert sheet["B2"].value == "01067" and sheet["B2"].number_format == "@"
    csv = CustomerExportService().write(frame, tmp_path / "export.csv", "csv")
    assert "01067" in csv.read_text(encoding="utf-8-sig") and "61462.0" not in csv.read_text(encoding="utf-8-sig")


def test_import_quality_warns_without_inventing_postal_code_and_saves_text(tmp_path):
    frame = pd.DataFrame({"KUNDENNAME": ["A", "B"], "PLZ": [1067, 61462.5]})
    service = ImportQualityService()
    analysis = service.analyze(frame, tmp_path / "source.xlsx")
    assert {issue.kind for issue in analysis.issues} >= {"short_postal_code", "invalid_postal_code"}
    result = service.unchanged(analysis)
    target = service.save_cleaned(result, tmp_path / "source.xlsx", tmp_path / "clean.xlsx")
    sheet = load_workbook(target)["Bereinigte Daten"]
    assert sheet["B2"].value == "1067" and sheet["B2"].number_format == "@"
    assert sheet["B3"].value == "61462.5"


def test_street_normalization_variants_and_conflicts():
    assert normalize_street(" Musterstraße 12 ") == normalize_street("Musterstr. 12")
    assert normalize_street("Musterstrasse 12") == normalize_street("Musterstr. 12")
    assert street_match("Musterstraße 12 A", "Musterstr. 12a") == "match"
    assert street_match("Musterstraße 12", "Musterstraße 14") == "conflict"
    assert street_match("Musterstraße 12", "Bahnhofstraße 12") == "conflict"


class MemoryDatabase:
    def __init__(self):
        self.row = None
        self.address = ("", "", "")

    def get_company(self, *_):
        return self.row

    def get_company_address(self, *_):
        return self.address

    def save_company(self, **values):
        self.address = (values.get("street", ""), values.get("zipcode", ""), values.get("country", ""))
        self.row = (1, values["company_name"], values["city"], values["website"], values["phone"],
                    values["email"], values["owner"], values["status"], values["source"], values["last_check"])


def service_with_candidates(candidates, contacts):
    service = ResearchService(MemoryDatabase())
    service.website_finder.ranked_candidates = lambda *_: [{"url": url} for url in candidates]
    service.contact_extractor.extract = lambda url: contacts[url]
    return service


def test_matching_abbreviated_address_accepts_contacts():
    service = service_with_candidates(["https://a.test"], {
        "https://a.test/": {"phone": "+49 30 123456", "email": "info@a.test",
                            "addresses": [{"street": "Musterstr. 12", "zipcode": "10115", "city": "Berlin"}]},
    })
    result = service.research("Restaurant", "Berlin", street="Musterstraße 12", zipcode="10115")
    assert result.website == "https://a.test/" and result.email == "info@a.test"


def test_wrong_first_candidate_is_rejected_and_second_used():
    contacts = {
        "https://wrong.test/": {"phone": "+49 30 999999", "email": "wrong@test.de",
                                "addresses": [{"street": "Bahnhofstraße 27", "city": "Berlin"}]},
        "https://right.test/": {"phone": "+49 30 123456", "email": "right@test.de",
                                "addresses": [{"street": "Hauptstr. 10", "city": "Berlin"}]},
    }
    result = service_with_candidates(list(contacts), contacts).research(
        "Restaurant", "Berlin", street="Hauptstraße 10", force_refresh=True)
    assert result.website == "https://right.test/" and result.email == "right@test.de"
    assert result.phone != "+49 30 999999"


def test_missing_or_conflicting_source_address_never_activates_result():
    for addresses in ([], [{"street": "Andere Straße 1", "city": "Berlin"}]):
        service = service_with_candidates(["https://wrong.test"], {
            "https://wrong.test/": {"phone": "+49 30 123456", "email": "info@test.de", "addresses": addresses},
        })
        result = service.research("Café", "Berlin", street="Hauptstraße 10")
        assert not result.website and not result.phone and not result.email
        assert result.status == "Nicht gefunden"


def test_optional_street_matching_controls_candidate_acceptance():
    contacts = {
        "https://candidate.test/": {
            "phone": "+49 30 123456",
            "email": "info@candidate.test",
            "addresses": [{"street": "Andere Straße 1", "city": "Berlin"}],
        }
    }
    strict = service_with_candidates(["https://candidate.test"], contacts).research(
        "Firma", "Berlin", street="Hauptstraße 10", use_street_matching=True
    )
    relaxed = service_with_candidates(["https://candidate.test"], contacts).research(
        "Firma", "Berlin", street="Hauptstraße 10", use_street_matching=False
    )
    assert strict.website == ""
    assert relaxed.website == "https://candidate.test/"
    assert relaxed.email == "info@candidate.test"


def test_bulk_forwards_clean_address_and_stable_id():
    captured = {}
    bulk = BulkResearchService()

    class FakeService:
        def research(self, company, city, **kwargs):
            captured.update(kwargs)
            return type("Result", (), {"company": company, "city": city, "status": "Nicht gefunden"})()

    bulk.research_service = FakeService()
    bulk.research_dataframe(pd.DataFrame([{
        "ID": 7, "KUNDENNAME": "Restaurant", "CITY": "Berlin", "STRASSE": "Hauptstr. 10",
        "ZIPCODE": float("nan"), "LAND": "DE",
    }]))
    assert captured == {
        "force_refresh": False,
        "street": "Hauptstr. 10",
        "zipcode": "",
        "country": "DE",
        "customer_id": 7,
        "use_street_matching": True,
    }


def test_bulk_forwards_disabled_street_matching():
    captured = {}
    bulk = BulkResearchService()

    class FakeService:
        def research(self, company, city, **kwargs):
            captured.update(kwargs)
            return type("Result", (), {"company": company, "city": city, "status": "Nicht gefunden"})()

    bulk.research_service = FakeService()
    bulk.research_dataframe(
        pd.DataFrame([{"KUNDENNAME": "Firma", "CITY": "Berlin", "STRASSE": "Hauptstraße 10"}]),
        use_street_matching=False,
    )
    assert captured["use_street_matching"] is False
