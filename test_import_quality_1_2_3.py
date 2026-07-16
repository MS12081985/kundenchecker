import math
import time

import pandas as pd
import pytest
from openpyxl import load_workbook

from config.app_config import AppConfig
from models.import_quality import ImportCleaningPlan
from services.import_quality_service import ImportQualityService, normalize_website
from excel.importer import load_excel


def frame(rows):
    return pd.DataFrame(rows, columns=("KUNDENNAME", "CITY", "STRASSE", "TELEFON", "EMAIL", "WEBSITE", "EXTRA"))


def test_identical_rows_normalize_case_spaces_nan_and_unicode():
    decomposed = "Mu\u0308ller GmbH"
    data = frame([
        {"KUNDENNAME": " Müller   GmbH ", "CITY": "Köln", "EXTRA": math.nan},
        {"KUNDENNAME": decomposed.lower(), "CITY": " KÖLN ", "EXTRA": None},
    ])
    analysis = ImportQualityService().analyze(data, "kunden.xlsx")
    assert analysis.identical_duplicates == 1
    assert analysis.duplicate_groups[0].category == "identical"


def test_complementary_same_company_and_city_is_safe_and_merged():
    data = frame([
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "TELEFON": "030 123456"},
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "EMAIL": "info@firma.de", "WEBSITE": "firma.de"},
    ])
    service = ImportQualityService()
    analysis = service.analyze(data)
    group = analysis.duplicate_groups[0]
    assert group.category == "safe"
    result = service.clean(analysis)
    assert len(result.dataframe) == 1
    assert result.dataframe.iloc[0]["EMAIL"] == "info@firma.de"
    assert result.report.merged_groups == 1


def test_identical_subset_and_complementary_row_form_one_safe_group():
    data = frame([
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "TELEFON": "030 123456"},
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "TELEFON": "030 123456"},
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "EMAIL": "info@firma.de"},
    ])
    analysis = ImportQualityService().analyze(data)
    assert len(analysis.duplicate_groups) == 1
    assert analysis.duplicate_groups[0].category == "safe"
    assert len(ImportQualityService().clean(analysis).dataframe) == 1


@pytest.mark.parametrize(
    ("field", "first", "second"),
    (
        ("CITY", "Berlin", "Hamburg"),
        ("TELEFON", "030 123456", "040 123456"),
        ("EMAIL", "a@firma.de", "b@firma.de"),
        ("WEBSITE", "https://a-firma.de", "https://b-firma.de"),
    ),
)
def test_conflicting_valid_values_are_never_automatic(field, first, second):
    rows = [
        {"KUNDENNAME": "Firma", "CITY": "Berlin", field: first},
        {"KUNDENNAME": "Firma", "CITY": "Berlin", field: second},
    ]
    analysis = ImportQualityService().analyze(frame(rows))
    group = analysis.duplicate_groups[0]
    assert group.category == "unsafe"
    assert not group.automatic
    assert len(ImportQualityService().clean(analysis).dataframe) == 2


def test_master_suggestion_and_manual_override():
    data = frame([
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "WEBSITE": "firma.de"},
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "TELEFON": "030 123456", "EMAIL": "info@firma.de"},
    ])
    service = ImportQualityService()
    analysis = service.analyze(data)
    group = analysis.duplicate_groups[0]
    second = analysis.rows[1]
    assert group.suggested_master_key == second.key
    first = analysis.rows[0]
    plan = ImportCleaningPlan(master_overrides={group.group_id: first.key})
    result = service.clean(analysis, plan)
    assert result.dataframe.iloc[0]["WEBSITE"] == "https://firma.de"


def test_missing_values_invalid_contacts_website_and_empty_row():
    data = frame([
        {"KUNDENNAME": "", "CITY": "", "TELEFON": "9 50189", "EMAIL": "bad@@mail", "WEBSITE": "firma.de"},
        {"KUNDENNAME": None, "CITY": None, "STRASSE": None, "TELEFON": None, "EMAIL": None, "WEBSITE": None, "EXTRA": None},
        {"KUNDENNAME": "Gültig", "CITY": "", "TELEFON": "+49 30 123456", "EMAIL": "info@gueltig.de"},
    ])
    analysis = ImportQualityService().analyze(data)
    assert analysis.missing_customer_name == 1
    assert analysis.missing_city == 2
    assert analysis.invalid_phones == 1
    assert analysis.invalid_emails == 1
    assert analysis.normalizable_websites == 1
    assert analysis.empty_rows == 1
    result = ImportQualityService().clean(analysis)
    assert list(result.dataframe["KUNDENNAME"]) == ["Gültig"]


def test_website_normalization_keeps_http_and_rejects_documents():
    assert normalize_website("example.de") == "https://example.de"
    assert normalize_website("http://example.de/") == "http://example.de"
    assert normalize_website("https://example.de/info.pdf") == ""


def test_similar_company_names_are_compared_only_as_candidates():
    data = frame([
        {"KUNDENNAME": "Mustertechnik GmbH", "CITY": "Berlin"},
        {"KUNDENNAME": "Mustertechnk GmbH", "CITY": "Berlin"},
    ])
    groups = ImportQualityService().analyze(data).duplicate_groups
    assert len(groups) == 1
    assert groups[0].category == "unsafe"


def test_missing_required_column_is_rejected():
    with pytest.raises(ValueError, match="KUNDENNAME"):
        ImportQualityService().analyze(pd.DataFrame({"CITY": ["Berlin"]}))


def test_cleaned_file_preserves_original_columns_and_writes_report(tmp_path):
    source = tmp_path / "Kunden.xlsx"
    original = frame([
        {"KUNDENNAME": "Firma", "CITY": "Berlin", "WEBSITE": "firma.de", "EXTRA": "bleibt"},
        {"KUNDENNAME": " Firma ", "CITY": "Berlin", "WEBSITE": "https://firma.de/", "EXTRA": "bleibt"},
    ])
    original.to_excel(source, index=False)
    before = source.read_bytes()
    service = ImportQualityService()
    result = service.clean(service.analyze(original, source))
    target = service.save_cleaned(result, source, tmp_path / "Kunden_bereinigt.xlsx")
    assert source.read_bytes() == before
    assert target != source
    workbook = load_workbook(target)
    assert workbook.sheetnames == ["Bereinigte Daten", "Importbericht"]
    assert [cell.value for cell in workbook["Bereinigte Daten"][1]] == list(original.columns)
    with pytest.raises(ValueError, match="Originaldatei"):
        service.save_cleaned(result, source, source)


def test_import_report_and_quality_score_are_transparent():
    data = frame([
        {"KUNDENNAME": "Komplett", "CITY": "Berlin", "TELEFON": "030 123456", "EMAIL": "info@komplett.de", "WEBSITE": "komplett.de"},
        {"KUNDENNAME": "Nur Name"},
    ])
    service = ImportQualityService()
    analysis = service.analyze(data)
    assert analysis.quality_score == round((100 + 40) / 2)
    report = service.clean(analysis).report
    assert report.rows_before == 2 and report.rows_after == 2


@pytest.mark.parametrize("size", (1_000, 10_000, 50_000))
def test_large_quality_scan_is_hash_based(size):
    data = pd.DataFrame({"KUNDENNAME": [f"Firma {index}" for index in range(size)], "CITY": "Berlin"})
    started = time.monotonic()
    metrics = ImportQualityService().dashboard_quality(data)
    assert metrics["duplicates"] == 0
    assert time.monotonic() - started < 20


def test_version_is_1_3_2():
    assert AppConfig.VERSION == "1.3.2"


@pytest.mark.parametrize(("filename", "engine"), (("kunden.xls", "xlrd"), ("kunden.xlsx", "openpyxl")))
def test_excel_import_uses_correct_engine(filename, engine, monkeypatch):
    calls = []
    monkeypatch.setattr(pd, "read_excel", lambda path, engine=None: calls.append((path, engine)) or pd.DataFrame({"KUNDENNAME": ["Firma"]}))
    assert len(load_excel(filename)) == 1
    assert calls == [(filename, engine)]


def test_excel_import_rejects_empty_data(monkeypatch):
    monkeypatch.setattr(pd, "read_excel", lambda *_args, **_kwargs: pd.DataFrame(columns=["KUNDENNAME"]))
    with pytest.raises(ValueError, match="keine Datenzeilen"):
        load_excel("leer.xlsx")
