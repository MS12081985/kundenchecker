"""Excel import analysis and cleaning without UI or database dependencies."""

from __future__ import annotations

import hashlib
import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, urlunparse

from loguru import logger

from models.import_quality import (
    ImportAnalysis,
    ImportCleaningPlan,
    ImportCleaningResult,
    ImportDuplicateGroup,
    ImportIssue,
    ImportReport,
    ImportRow,
)
from services.contact_validator import validate_email, validate_phone_details
from services.website_finder import WebsiteFinder
from rapidfuzz import fuzz
from models.address_utils import POSTAL_CODE_COLUMNS, normalize_postal_code, normalize_street


NAME = "KUNDENNAME"
CITY = "CITY"
PHONE_COLUMNS = ("TELEFON", "PHONE")
EMAIL_COLUMNS = ("EMAIL", "E-MAIL")
WEBSITE_COLUMNS = ("WEBSITE", "URL")
ADDRESS_COLUMNS = ("STRASSE", "STREET", "ADRESSE", "ADDRESS")


def _blank(value):
    return value is None or (isinstance(value, float) and math.isnan(value)) or not str(value).strip()


def _text(value):
    return "" if _blank(value) else re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value))).strip()


def _name(value):
    value = _text(value).casefold()
    return re.sub(r"[^\w]+", " ", value).strip()


def _email(value):
    return validate_email(value)


def _phone(value):
    return validate_phone_details(value).value


def normalize_website(value):
    raw = _text(value)
    if not raw:
        return ""
    if not re.match(r"^[a-zA-Z][a-zA-Z0-9+.-]*://", raw):
        raw = "https://" + raw
    parsed = urlparse(raw)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return ""
    if any(parsed.path.lower().endswith(extension) for extension in WebsiteFinder.DOCUMENT_EXTENSIONS):
        return ""
    path = parsed.path or "/"
    cleaned = WebsiteFinder.clean_url(
        urlunparse((parsed.scheme.lower(), parsed.netloc.lower(), path, "", parsed.query, ""))
    )
    return cleaned.rstrip("/") if cleaned else ""


class ImportQualityService:
    def dashboard_quality(self, dataframe):
        import pandas as pd

        if dataframe is None or dataframe.empty:
            return {"quality_score": 0, "invalid_phone": 0, "invalid_email": 0, "duplicates": 0}
        seen = set()
        duplicates = invalid_phone = invalid_email = points = 0
        for _, row in dataframe.iterrows():
            name, city = _name(row.get(NAME)), _name(row.get(CITY))
            phone_value = self._first(row, PHONE_COLUMNS)
            email_value = self._first(row, EMAIL_COLUMNS)
            website = normalize_website(self._first(row, WEBSITE_COLUMNS))
            phone, email = _phone(phone_value), _email(email_value)
            invalid_phone += int(not _blank(phone_value) and not phone)
            invalid_email += int(not _blank(email_value) and not email)
            key = (name, city)
            duplicate = bool(name and key in seen)
            duplicates += int(duplicate)
            seen.add(key)
            points += 35 * bool(name) + 10 * bool(city) + 20 * bool(website)
            points += 15 * bool(phone) + 15 * bool(email) + 5 * (not duplicate)
        return {"quality_score": round(points / len(dataframe)), "invalid_phone": invalid_phone,
                "invalid_email": invalid_email, "duplicates": duplicates}

    def analyze(self, dataframe, source_path="", sheet_name="Tabelle1"):
        import pandas as pd

        if dataframe is None:
            dataframe = pd.DataFrame()
        if NAME not in dataframe.columns:
            raise ValueError("Die Excel-Datei enthält die erforderliche Spalte KUNDENNAME nicht.")
        logger.info("Importanalyse gestartet: Datei={} Sheet={}", Path(source_path).name, sheet_name)
        rows = []
        issues = []
        normalized = {}
        for position, (index, series) in enumerate(dataframe.iterrows(), start=2):
            values = series.to_dict()
            key = hashlib.sha256(f"{Path(source_path).resolve()}|{sheet_name}|{position}".encode()).hexdigest()[:20]
            row = ImportRow(key, position, index, values)
            rows.append(row)
            item = self._normalized_row(values, dataframe.columns)
            normalized[key] = item
            if item["empty"]:
                issues.append(ImportIssue(key, position, "empty_row", "", "Vollständig leere Zeile"))
                continue
            if not item["name"]:
                issues.append(ImportIssue(key, position, "missing", NAME, "Kundenname fehlt"))
            if not item["city"]:
                issues.append(ImportIssue(key, position, "missing", CITY, "Ort fehlt"))
            phone_value = self._first(values, PHONE_COLUMNS)
            if not _blank(phone_value) and not item["phone"]:
                issues.append(ImportIssue(key, position, "invalid_phone", "TELEFON", "Ungültige Telefonnummer", _text(phone_value)))
            email_value = self._first(values, EMAIL_COLUMNS)
            if not _blank(email_value) and not item["email"]:
                issues.append(ImportIssue(key, position, "invalid_email", "EMAIL", "Ungültige E-Mail-Adresse", _text(email_value)))
            website_value = self._first(values, WEBSITE_COLUMNS)
            if not _blank(website_value) and item["website"] and item["website"] != _text(website_value).rstrip("/"):
                issues.append(ImportIssue(key, position, "website_normalizable", "WEBSITE", "Website kann normalisiert werden", _text(website_value)))
            for postal_column in (column for column in dataframe.columns if str(column).upper() in POSTAL_CODE_COLUMNS):
                raw_postal = _text(values.get(postal_column))
                postal = normalize_postal_code(values.get(postal_column))
                if re.fullmatch(r"[+-]?\d+[.,]\d+", raw_postal) and re.fullmatch(r".*[.,]0+", raw_postal) is None:
                    issues.append(ImportIssue(key, position, "invalid_postal_code", str(postal_column),
                                              "Postleitzahl enthält einen Dezimalwert", raw_postal))
                elif postal.isdigit() and len(postal) < 5:
                    issues.append(ImportIssue(key, position, "short_postal_code", str(postal_column),
                                              "Kurze Postleitzahl; eine verlorene führende Null wird nicht ergänzt", postal))

        groups = self._duplicate_groups(rows, normalized, tuple(dataframe.columns))
        counts = lambda kind: sum(issue.kind == kind for issue in issues)
        nonempty = sum(not normalized[row.key]["empty"] for row in rows)
        missing_names = sum(issue.kind == "missing" and issue.field == NAME for issue in issues)
        score = self.quality_score(rows, normalized, groups)
        analysis = ImportAnalysis(
            str(source_path), sheet_name, datetime.now().isoformat(timespec="seconds"), dataframe.copy(),
            tuple(rows), tuple(issues), tuple(groups), len(rows), max(0, nonempty - missing_names),
            sum(len(group.row_keys) - 1 for group in groups if group.category == "identical"),
            sum(group.category != "identical" for group in groups), missing_names,
            sum(issue.kind == "missing" and issue.field == CITY for issue in issues),
            counts("invalid_phone"), counts("invalid_email"),
            sum(not normalized[row.key]["website"] for row in rows if not normalized[row.key]["empty"]),
            counts("website_normalizable"), counts("empty_row"), score,
        )
        logger.info("Importanalyse abgeschlossen: Zeilen={} Gruppen={}", len(rows), len(groups))
        return analysis

    @staticmethod
    def _first(values, names):
        for name in names:
            if name in values:
                return values[name]
        return ""

    def _normalized_row(self, values, columns):
        phone = _phone(self._first(values, PHONE_COLUMNS))
        email = _email(self._first(values, EMAIL_COLUMNS))
        website = normalize_website(self._first(values, WEBSITE_COLUMNS))
        address = normalize_street(self._first(values, ADDRESS_COLUMNS)).name
        all_values = tuple(self._normalize_field(column, values.get(column)) for column in columns)
        return {
            "name": _name(values.get(NAME)), "city": _name(values.get(CITY)), "phone": phone,
            "email": email, "website": website, "address": address, "all": all_values,
            "empty": all(_blank(value) for value in values.values()),
        }

    def _normalize_field(self, column, value):
        if column in PHONE_COLUMNS:
            return _phone(value) or _text(value).casefold()
        if column in EMAIL_COLUMNS:
            return _email(value) or _text(value).casefold()
        if column in WEBSITE_COLUMNS:
            return normalize_website(value) or _text(value).casefold()
        if str(column).upper() in POSTAL_CODE_COLUMNS:
            return normalize_postal_code(value)
        return _name(value)

    def _duplicate_groups(self, rows, normalized, columns):
        groups = []
        assigned = set()
        by_name = defaultdict(list)
        for row in rows:
            item = normalized[row.key]
            if not item["empty"] and item["name"]:
                by_name[item["name"]].append(row)
        for members in by_name.values():
            if len(members) < 2:
                continue
            if len({normalized[row.key]["all"] for row in members}) == 1:
                conflicts = ()
                category = "identical"
            else:
                conflicts = self._conflicts(members, normalized)
                cities = {normalized[row.key]["city"] for row in members} - {""}
                evidence = self._shared_identity(members, normalized)
                category = "safe" if len(cities) <= 1 and evidence and not conflicts else "unsafe"
            groups.append(self._group(members, category, normalized, columns, conflicts))
            assigned.update(row.key for row in members)

        # Candidate signatures avoid comparing every row with every other row.
        candidates = [row for row in rows if row.key not in assigned and normalized[row.key]["name"]]
        parent = {row.key: row.key for row in candidates}
        lookup = {}

        def root(key):
            while parent[key] != key:
                parent[key] = parent[parent[key]]
                key = parent[key]
            return key

        for row in candidates:
            item = normalized[row.key]
            compact = item["name"].replace(" ", "")
            signatures = {compact}
            signatures.update(compact[:index] + compact[index + 1 :] for index in range(len(compact)))
            for signature in signatures:
                candidate = lookup.get((item["city"], signature))
                candidate_name = normalized[candidate.key]["name"] if candidate else ""
                numbers = re.findall(r"\d+", item["name"])
                candidate_numbers = re.findall(r"\d+", candidate_name)
                compatible_numbers = not (numbers or candidate_numbers) or numbers == candidate_numbers
                if candidate and compatible_numbers and fuzz.ratio(item["name"], candidate_name) >= 90:
                    parent[root(row.key)] = root(candidate.key)
                else:
                    lookup[(item["city"], signature)] = row
        fuzzy_groups = defaultdict(list)
        for row in candidates:
            fuzzy_groups[root(row.key)].append(row)
        for members in fuzzy_groups.values():
            if len(members) > 1:
                conflicts = self._conflicts(members, normalized)
                groups.append(self._group(members, "unsafe", normalized, columns, conflicts))
        return groups

    def _group(self, members, category, normalized, columns, conflicts=()):
        master = min(members, key=lambda row: self._master_rank(row, normalized[row.key]))
        differences = tuple(
            column for column in columns
            if len({_text(row.values.get(column)) for row in members} - {""}) > 1
        )
        group_id = hashlib.sha256("|".join(row.key for row in members).encode()).hexdigest()[:16]
        return ImportDuplicateGroup(group_id, category, tuple(row.key for row in members),
                                    tuple(row.excel_row for row in members), master.key,
                                    tuple(conflicts), differences)

    @staticmethod
    def _shared_identity(members, normalized):
        for field in ("website", "address"):
            values = [normalized[row.key][field] for row in members]
            if values[0] and len(set(values)) == 1:
                return True
        return True  # same normalized company and city with complementary data

    @staticmethod
    def _conflicts(members, normalized):
        conflicts = []
        for field in ("city", "address", "phone", "email", "website"):
            values = {normalized[row.key][field] for row in members} - {""}
            if len(values) > 1:
                conflicts.append(field)
        return tuple(conflicts)

    @staticmethod
    def _master_rank(row, item):
        both = bool(item["phone"] and item["email"])
        filled = sum(not _blank(value) for value in row.values.values())
        return (-int(both), -int(bool(item["website"])), -int(bool(item["address"])), -filled, row.excel_row)

    @staticmethod
    def quality_score(rows, normalized, groups):
        active = [row for row in rows if not normalized[row.key]["empty"]]
        if not active:
            return 0
        duplicate_keys = {key for group in groups for key in group.row_keys[1:]}
        points = 0
        for row in active:
            item = normalized[row.key]
            points += 35 * bool(item["name"])
            points += 10 * bool(item["city"])
            points += 20 * bool(item["website"])
            points += 15 * bool(item["phone"])
            points += 15 * bool(item["email"])
            points += 5 * (row.key not in duplicate_keys)
        return round(points / len(active))

    def clean(self, analysis, plan=None):
        plan = plan or ImportCleaningPlan()
        frame = analysis.dataframe.copy()
        row_by_key = {row.key: row for row in analysis.rows}
        drop = set()
        removed_identical = merged_groups = 0
        for group in analysis.duplicate_groups:
            if group.category == "identical" and plan.remove_identical:
                should_merge = True
            elif group.category == "safe" and plan.merge_safe and not group.conflicts:
                should_merge = True
            else:
                continue
            master_key = plan.master_overrides.get(group.group_id, group.suggested_master_key)
            if master_key not in group.row_keys:
                raise ValueError("Die gewählte Hauptzeile gehört nicht zur Dublettengruppe.")
            master = row_by_key[master_key]
            for key in group.row_keys:
                if key == master_key:
                    continue
                other = row_by_key[key]
                for column in frame.columns:
                    if _blank(frame.at[master.source_index, column]) and not _blank(frame.at[other.source_index, column]):
                        frame.at[master.source_index, column] = frame.at[other.source_index, column]
                drop.add(other.source_index)
            if group.category == "identical":
                removed_identical += len(group.row_keys) - 1
            else:
                merged_groups += 1

        skipped = corrected_phones = discarded_emails = normalized_websites = 0
        for row in analysis.rows:
            if row.source_index in drop:
                continue
            if plan.skip_missing_customer_name and _blank(frame.at[row.source_index, NAME]):
                drop.add(row.source_index); skipped += 1; continue
            for column in PHONE_COLUMNS:
                if column in frame.columns and not _blank(frame.at[row.source_index, column]):
                    valid = _phone(frame.at[row.source_index, column])
                    if valid and valid != _text(frame.at[row.source_index, column]):
                        frame.at[row.source_index, column] = valid; corrected_phones += 1
                    elif not valid and plan.invalid_phone_action == "clear":
                        frame.at[row.source_index, column] = ""; corrected_phones += 1
            for column in EMAIL_COLUMNS:
                if column in frame.columns and not _blank(frame.at[row.source_index, column]):
                    valid = _email(frame.at[row.source_index, column])
                    if valid:
                        frame.at[row.source_index, column] = valid
                    elif plan.invalid_email_action == "clear":
                        frame.at[row.source_index, column] = ""; discarded_emails += 1
            for column in WEBSITE_COLUMNS:
                if column in frame.columns and not _blank(frame.at[row.source_index, column]):
                    valid = normalize_website(frame.at[row.source_index, column])
                    if valid and valid != _text(frame.at[row.source_index, column]).rstrip("/"):
                        frame.at[row.source_index, column] = valid; normalized_websites += 1
            for column in frame.columns:
                if str(column).upper() in POSTAL_CODE_COLUMNS:
                    frame.at[row.source_index, column] = normalize_postal_code(frame.at[row.source_index, column])
        frame = frame.drop(index=list(drop)).reset_index(drop=True)
        open_conflicts = sum(bool(group.conflicts) for group in analysis.duplicate_groups)
        report = ImportReport(analysis.source_path, analysis.analyzed_at, analysis.total_rows, len(frame), len(frame),
                              removed_identical, merged_groups, skipped, corrected_phones,
                              discarded_emails, normalized_websites, open_conflicts)
        logger.info("Importbereinigung abgeschlossen: vorher={} nachher={} Gruppen={}", analysis.total_rows, len(frame), merged_groups)
        return ImportCleaningResult(frame, report)

    def unchanged(self, analysis):
        report = ImportReport(analysis.source_path, analysis.analyzed_at, analysis.total_rows,
                              len(analysis.dataframe), len(analysis.dataframe), 0, 0, 0, 0, 0, 0,
                              sum(bool(group.conflicts) for group in analysis.duplicate_groups))
        return ImportCleaningResult(analysis.dataframe.copy(), report)

    def save_cleaned(self, result, source_path, target_path):
        import pandas as pd
        from openpyxl import load_workbook

        source = Path(source_path).expanduser().resolve()
        target = Path(target_path).expanduser()
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        if target.resolve() == source:
            raise ValueError("Die Originaldatei darf nicht überschrieben werden.")
        target.parent.mkdir(parents=True, exist_ok=True)
        report = result.report
        export = result.dataframe.copy()
        postal_columns = [column for column in export.columns if str(column).upper() in POSTAL_CODE_COLUMNS]
        for column in postal_columns:
            export[column] = export[column].map(normalize_postal_code)
        report_rows = [
            ("Quelldatei", Path(report.source_path).name), ("Analysezeitpunkt", report.analyzed_at),
            ("Zeilen vorher", report.rows_before), ("Zeilen nachher", report.rows_after),
            ("Entfernte identische Dubletten", report.removed_identical),
            ("Zusammengeführte Gruppen", report.merged_groups), ("Übersprungene Zeilen", report.skipped_rows),
            ("Ungültige Telefonnummern korrigiert", report.corrected_phones),
            ("Ungültige E-Mails verworfen", report.discarded_emails),
            ("Normalisierte Websites", report.normalized_websites), ("Offene Konflikte", report.open_conflicts),
        ]
        try:
            with pd.ExcelWriter(target, engine="openpyxl") as writer:
                export.to_excel(writer, index=False, sheet_name="Bereinigte Daten")
                pd.DataFrame(report_rows, columns=("Kennzahl", "Wert")).to_excel(writer, index=False, sheet_name="Importbericht")
        except OSError as error:
            logger.exception("Bereinigte Excel-Datei konnte nicht gespeichert werden")
            raise ValueError("Die bereinigte Excel-Datei konnte nicht gespeichert werden.") from error
        workbook = load_workbook(target)
        sheet = workbook["Bereinigte Daten"]
        for column in postal_columns:
            position = list(export.columns).index(column) + 1
            for cells in sheet.iter_cols(min_col=position, max_col=position, min_row=2):
                for cell in cells:
                    cell.number_format = "@"
        workbook.save(target)
        logger.info("Bereinigte Excel-Datei gespeichert: {}", target)
        return target
