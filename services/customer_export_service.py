"""Customer export selection and file writing without UI dependencies."""

from __future__ import annotations

from pathlib import Path
from models.address_utils import POSTAL_CODE_COLUMNS, normalize_postal_code
from models.customer_status import normalize_customer_status, status_mask

EXPORT_SCOPES = {
    "visible": None,
    "complete": "complete",
    "active": "active",
    "contactable": "complete_or_active",
    "inactive": "inactive",
    "not_found": "not_found",
    "selected": None,
    "all_loaded": None,
}
STANDARD_COLUMNS = (
    "KUNDENNAME", "CITY", "PLZ", "POSTLEITZAHL", "STRASSE", "STREET",
    "LAND", "COUNTRY", "TELEFON", "EMAIL", "WEBSITE", "STATUS",
)
CRM_COLUMNS = (
    "ANSPRECHPARTNER", "POSITION", "DIREKTTELEFON", "DIREKTE_EMAIL",
    "KUNDENSTATUS", "PRIORITÄT", "TAGS", "NOTIZEN", "LETZTER_KONTAKT",
    "NÄCHSTE_WIEDERVORLAGE",
)
ENRICHMENT_COLUMNS = (
    "WEBSITE_SCORE", "WEBSITE_SCORE_CATEGORY", "INDUSTRY", "INDUSTRY_CONFIDENCE",
    "OPENING_HOURS", "SOCIAL_FACEBOOK", "SOCIAL_INSTAGRAM", "SOCIAL_LINKEDIN",
    "SOCIAL_YOUTUBE", "SOCIAL_TIKTOK", "SOCIAL_X", "SOCIAL_PINTEREST",
    "HAS_IMPRINT", "IMPRINT_URL", "HAS_PRIVACY_POLICY", "PRIVACY_URL",
    "CONTACT_FORM_URL", "SHORT_DESCRIPTION", "ANALYZED_AT", "ENRICHMENT_STATUS", "ENRICHMENT_ERROR",
    "IMPRINT_OWNER_NAMES", "IMPRINT_MANAGING_DIRECTOR_NAMES", "IMPRINT_REPRESENTATIVE_NAMES",
    "IMPRINT_LEGAL_FORM", "IMPRINT_COMPANY_NAME", "IMPRINT_STREET", "IMPRINT_HOUSE_NUMBER", "IMPRINT_POSTAL_CODE",
    "IMPRINT_CITY", "IMPRINT_COUNTRY", "IMPRINT_PHONE", "IMPRINT_EMAIL", "IMPRINT_VAT_ID",
    "IMPRINT_REGISTER_TYPE", "IMPRINT_REGISTER_NUMBER", "IMPRINT_REGISTER_COURT",
    "IMPRINT_CONFIDENCE", "IMPRINT_ANALYZED_AT",
)


normalize_status = normalize_customer_status


class CustomerExportService:
    @staticmethod
    def select(dataframe, scope="visible", selected_keys=None):
        import pandas as pd
        frame = dataframe.copy() if dataframe is not None else pd.DataFrame()
        if frame.empty:
            return frame
        if scope not in EXPORT_SCOPES:
            raise ValueError("Unbekannter Exportumfang")
        if scope == "selected":
            keys = {(str(name), str(city)) for name, city in (selected_keys or set())}
            if not keys:
                return frame.iloc[0:0].copy()
            mask = frame.apply(lambda row: (str(row.get("KUNDENNAME", "")), str(row.get("CITY", ""))) in keys, axis=1)
            return frame.loc[mask].reset_index(drop=True)
        status_filter = EXPORT_SCOPES[scope]
        if status_filter is not None:
            if "STATUS" not in frame.columns:
                return frame.iloc[0:0].copy()
            mask = status_mask(frame["STATUS"], status_filter)
            return frame.loc[mask].reset_index(drop=True)
        return frame.reset_index(drop=True)

    @staticmethod
    def columns(dataframe, *, include_crm=True, include_enrichment=True, include_technical=False):
        if include_technical:
            return dataframe.copy()
        allowed = list(STANDARD_COLUMNS) + (list(CRM_COLUMNS) if include_crm else []) + (list(ENRICHMENT_COLUMNS) if include_enrichment else [])
        return dataframe.loc[:, [column for column in allowed if column in dataframe.columns]].copy()

    @staticmethod
    def target_path(filename, export_format):
        path = Path(filename)
        expected = ".csv" if export_format == "csv" else ".xlsx"
        if path.suffix and path.suffix.lower() != expected:
            raise ValueError(f"Das gewählte Format benötigt die Dateiendung {expected}.")
        return path if path.suffix else path.with_suffix(expected)

    def write(self, dataframe, filename, export_format):
        path = self.target_path(filename, export_format)
        if path.suffix.lower() not in {".xlsx", ".csv"}:
            raise ValueError("Bitte wählen Sie eine .xlsx- oder .csv-Datei.")
        export = dataframe.copy()
        postal_columns = [column for column in export.columns if str(column).upper() in POSTAL_CODE_COLUMNS]
        for column in postal_columns:
            export[column] = export[column].map(normalize_postal_code)
        if path.suffix.lower() == ".csv":
            export.to_csv(path, index=False, encoding="utf-8-sig")
            return path
        export.to_excel(path, index=False, engine="openpyxl")
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        workbook = load_workbook(path); sheet = workbook.active
        fill = PatternFill("solid", fgColor="2F5597")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF"); cell.fill = fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in postal_columns:
            position = list(export.columns).index(column) + 1
            for cell in sheet.iter_cols(min_col=position, max_col=position, min_row=2):
                for value_cell in cell:
                    value_cell.number_format = "@"
        for column in sheet.columns:
            width = min(60, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        workbook.save(path)
        return path
